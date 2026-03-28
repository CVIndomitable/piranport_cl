#!/usr/bin/env python3
"""生成皮兰港模组物品数值总表 Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 颜色常量 ──────────────────────────────────────────────────────────────
H_FILL   = PatternFill("solid", fgColor="2F4F8F")   # 深蓝表头
H_FONT   = Font(bold=True, color="FFFFFF", size=10)
S1_FILL  = PatternFill("solid", fgColor="D9E1F2")   # 浅蓝分组
S2_FILL  = PatternFill("solid", fgColor="EBF0FA")
NORM_FONT = Font(size=10)
CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN     = Side(style="thin", color="AAAAAA")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def header(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.fill  = H_FILL
        cell.font  = H_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_rows(ws, start=2, alt=True):
    for i, row in enumerate(ws.iter_rows(min_row=start)):
        fill = S1_FILL if (alt and i % 2 == 0) else S2_FILL
        for cell in row:
            cell.fill = fill
            cell.font = NORM_FONT
            cell.border = BORDER
            cell.alignment = CENTER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ══════════════════════════════════════════════════════════════════════════
# Sheet 1: 物品总表
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "物品总表"
header(ws1, ["注册名", "中文名", "分类", "备注"])

items_all = [
    # 矿物/材料
    ("bauxite_ore",           "矾土矿",         "矿物/材料", "世界生成，可熔炼/冶炼"),
    ("aluminum_block",        "铝块",           "矿物/材料", "9铝锭合成"),
    ("salt_block",            "盐块",           "矿物/材料", "河床生成，9盐合成"),
    ("aluminum_ingot",        "铝锭",           "矿物/材料", "矾土矿冶炼"),
    ("salt",                  "盐",             "矿物/材料", "盐块破碎"),
    # 舰装核心
    ("small_ship_core",       "小型舰装核心",   "舰装核心", "2武器槽/4弹药槽/1强化槽，负重40"),
    ("medium_ship_core",      "中型舰装核心",   "舰装核心", "4武器槽/6弹药槽/2强化槽，负重72"),
    ("large_ship_core",       "大型舰装核心",   "舰装核心", "6武器槽/8弹药槽/3强化槽，负重112"),
    # 火炮
    ("small_gun",             "小型火炮",       "火炮",     "负重6"),
    ("medium_gun",            "中型火炮",       "火炮",     "负重16"),
    ("large_gun",             "大型火炮",       "火炮",     "负重30"),
    # 炮弹
    ("small_he_shell",        "小口径HE炮弹",   "炮弹",     "范围爆炸"),
    ("medium_he_shell",       "中口径HE炮弹",   "炮弹",     "范围爆炸"),
    ("large_he_shell",        "大口径HE炮弹",   "炮弹",     "范围爆炸"),
    ("small_ap_shell",        "小口径AP炮弹",   "炮弹",     "穿甲，×1.3伤害"),
    ("medium_ap_shell",       "中口径AP炮弹",   "炮弹",     "穿甲，×1.3伤害"),
    ("large_ap_shell",        "大口径AP炮弹",   "炮弹",     "穿甲，×1.3伤害"),
    # 鱼雷
    ("torpedo_533mm",         "533mm标准鱼雷",  "鱼雷",     "stacksTo 16"),
    ("torpedo_610mm",         "610mm氧气鱼雷",  "鱼雷",     "stacksTo 16"),
    ("twin_torpedo_launcher", "双联装鱼雷发射器","鱼雷发射器","533mm，2联±3°，耐久64，负重8"),
    ("triple_torpedo_launcher","三联装鱼雷发射器","鱼雷发射器","533mm，3联±4°，耐久48，负重12"),
    ("quad_torpedo_launcher", "四联装鱼雷发射器","鱼雷发射器","610mm，4联，耐久32，负重20"),
    # 装甲板
    ("small_armor_plate",     "小型附加装甲",   "装甲",     "护甲+2，负重10"),
    ("medium_armor_plate",    "中型附加装甲",   "装甲",     "护甲+4，负重20"),
    ("large_armor_plate",     "大型附加装甲",   "装甲",     "护甲+6，负重30"),
    # 飞机
    ("fighter_squadron",      "战斗机编队",     "飞机",     "见航空系统表"),
    ("dive_bomber_squadron",  "俯冲轰炸机编队", "飞机",     "见航空系统表"),
    ("torpedo_bomber_squadron","鱼雷轰炸机编队","飞机",     "见航空系统表"),
    ("level_bomber_squadron", "水平轰炸机编队", "飞机",     "见航空系统表"),
    # 航空弹药
    ("aviation_fuel",         "航空燃料",       "航空弹药", "变身时自动补充飞机燃料"),
    ("aerial_bomb",           "航空炸弹",       "航空弹药", "俯冲/水平轰炸机弹药"),
    ("aerial_torpedo",        "航空鱼雷",       "航空弹药", "鱼雷轰炸机弹药"),
    ("fighter_ammo",          "子弹",           "航空弹药", "战斗机弹药，64发"),
    # 农作物
    ("tomato_seeds",          "番茄种子",       "种子",     ""),
    ("soybean_seeds",         "大豆种子",       "种子",     ""),
    ("chili_seeds",           "辣椒种子",       "种子",     ""),
    ("onion_seeds",           "洋葱种子",       "种子",     ""),
    ("rice_seeds",            "稻种",           "种子",     "需水淹土壤"),
    ("lettuce_seeds",         "生菜种子",       "种子",     ""),
    ("garlic_seeds",          "大蒜种子",       "种子",     ""),
    ("tomato",                "番茄",           "作物",     ""),
    ("soybean",               "大豆",           "作物",     ""),
    ("chili",                 "辣椒",           "作物",     ""),
    ("onion",                 "洋葱",           "作物",     ""),
    ("rice",                  "米",             "作物",     ""),
    ("lettuce",               "生菜",           "作物",     ""),
    ("garlic",                "大蒜",           "作物",     ""),
    # 食材/调料
    ("flour",                 "面粉",           "食材",     "石磨：小麦"),
    ("rice_flour",            "米粉",           "食材",     "石磨：米"),
    ("chili_powder",          "辣椒粉",         "食材",     "石磨：辣椒+盐"),
    ("pork_paste",            "猪肉糜",         "食材",     "石磨：生猪排"),
    ("edible_oil",            "食用油",         "食材",     "工作台"),
    ("butter",                "黄油",           "食材",     "工作台"),
    ("cream",                 "奶油",           "食材",     "工作台"),
    ("soybean_milk",          "豆浆",           "食材",     "工作台"),
    ("tofu",                  "豆腐",           "食材",     "工作台"),
    ("cheese",                "奶酪",           "食材",     "工作台"),
    ("yeast",                 "酵母瓶",         "食材",     "酿造台"),
    ("soy_sauce",             "酱油",           "食材",     "酿造台"),
    ("vinegar",               "醋",             "食材",     "酿造台"),
    ("cooking_wine",          "料酒",           "食材",     "酿造台"),
    ("miso",                  "味噌",           "食材",     "酿造台"),
    ("brine",                 "盐水",           "食材",     "酿造台"),
    ("pie_crust",             "馅饼酥皮",       "食材",     "工作台"),
    ("raw_pasta",             "生意面",         "食材",     "工作台"),
    ("fermented_fish",        "发酵鱼",         "食材",     "工作台"),
    ("pizza_base",            "披萨饼底",       "食材",     "工作台"),
    ("gypsum_chip",           "石膏碎片",       "食材",     "工作台"),
    ("quicklime",             "生石灰",         "食材",     "熔炉：骨头"),
    # 中间产物
    ("sausage",               "香肠",           "中间产物", ""),
    ("sliced_sausage",        "切片香肠",       "中间产物", "砧板：香肠"),
    ("bacon",                 "培根",           "中间产物", "砧板：熟猪排"),
    ("toast_bread_slices",    "吐司面包片",     "中间产物", "砧板：面包"),
    ("beer",                  "啤酒",           "中间产物", "酿造台"),
    ("round_bun",             "圆面包",         "中间产物", ""),
    # 食物
    ("toast_bread",           "吐司面包",       "食物",     "见食物数值表"),
    ("naval_baked_beans",     "海军烘豆子",     "食物",     "见食物数值表"),
    ("latiao",                "辣条",           "食物",     "见食物数值表"),
    ("mapo_tofu",             "麻婆豆腐",       "食物",     "见食物数值表"),
    ("naval_curry",           "海军咖喱",       "食物",     "见食物数值表"),
    ("fried_fish_and_chips",  "炸鱼薯条",       "食物",     "见食物数值表"),
    ("scone",                 "司康饼",         "食物",     "见食物数值表"),
    ("salted_egg_tofu",       "咸蛋拌豆腐",     "食物",     "见食物数值表"),
    ("surstromming",          "鲱鱼罐头",       "食物",     "见食物数值表"),
    ("american_burger",       "美式汉堡",       "食物",     "见食物数值表"),
    ("hotdog",                "热狗",           "食物",     "见食物数值表"),
    ("pasta",                 "意面",           "食物",     "见食物数值表"),
    ("cooked_rice",           "米饭",           "食物",     "见食物数值表"),
    ("beet_blossom",          "甜豆花",         "食物",     "见食物数值表"),
    ("miso_soup",             "味噌汤",         "食物",     "见食物数值表"),
    # 功能方块
    ("stone_mill",            "石磨",           "功能方块", "研磨加工站"),
    ("cutting_board",         "砧板",           "功能方块", "切割加工站"),
    ("cooking_pot",           "厨锅",           "功能方块", "烹饪加工站"),
    # 其他
    ("floating_target",       "浮动靶子",       "测试工具", "水面漂浮，可穿盔甲"),
    ("guidebook",             "航行手册",       "书籍",     "Patchouli软依赖"),
]

for row in items_all:
    ws1.append(["piranport:" + row[0]] + list(row[1:]))

style_rows(ws1)
set_col_widths(ws1, [32, 18, 12, 28])
freeze(ws1)

# ══════════════════════════════════════════════════════════════════════════
# Sheet 2: 战斗数值
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("战斗数值")
header(ws2, ["注册名", "中文名", "分类", "负重", "伤害/护甲", "弹药口径/散布", "耐久", "备注"])

combat_data = [
    # 舰装核心
    ("small_ship_core",  "小型舰装核心", "核心", 0, "-",  "2武器/4弹药/1强化", "-",  "最大负重40"),
    ("medium_ship_core", "中型舰装核心", "核心", 0, "-",  "4武器/6弹药/2强化", "-",  "最大负重72"),
    ("large_ship_core",  "大型舰装核心", "核心", 0, "-",  "6武器/8弹药/3强化", "-",  "最大负重112"),
    # 火炮
    ("small_gun",   "小型火炮", "火炮", 6,  "伤害×1(HE爆炸/AP×1.3)", "小口径",   "-",  ""),
    ("medium_gun",  "中型火炮", "火炮", 16, "伤害×2",                "中口径",   "-",  ""),
    ("large_gun",   "大型火炮", "火炮", 30, "伤害×3",                "大口径",   "-",  ""),
    # 炮弹
    ("small_he_shell",  "小口径HE炮弹", "炮弹", 0, "HE爆炸", "小口径", "-", "×4产出"),
    ("medium_he_shell", "中口径HE炮弹", "炮弹", 0, "HE爆炸", "中口径", "-", "×2产出"),
    ("large_he_shell",  "大口径HE炮弹", "炮弹", 0, "HE爆炸", "大口径", "-", "×1产出"),
    ("small_ap_shell",  "小口径AP炮弹", "炮弹", 0, "AP穿甲×1.3", "小口径", "-", "×4产出"),
    ("medium_ap_shell", "中口径AP炮弹", "炮弹", 0, "AP穿甲×1.3", "中口径", "-", "×2产出"),
    ("large_ap_shell",  "大口径AP炮弹", "炮弹", 0, "AP穿甲×1.3", "大口径", "-", "×1产出"),
    # 鱼雷/发射器
    ("torpedo_533mm",          "533mm标准鱼雷",   "鱼雷",     0,  "-",        "533mm", "-",  "stacksTo 16"),
    ("torpedo_610mm",          "610mm氧气鱼雷",   "鱼雷",     0,  "-",        "610mm", "-",  "stacksTo 16"),
    ("twin_torpedo_launcher",  "双联装鱼雷发射器","鱼雷发射器", 8, "进水debuff", "533mm ±3°×2", 64, ""),
    ("triple_torpedo_launcher","三联装鱼雷发射器","鱼雷发射器",12, "进水debuff", "533mm ±4°×3", 48, ""),
    ("quad_torpedo_launcher",  "四联装鱼雷发射器","鱼雷发射器",20, "进水debuff", "610mm ×4",    32, ""),
    # 装甲板
    ("small_armor_plate",  "小型附加装甲", "装甲", 10, "+2护甲", "-", "-", "强化槽"),
    ("medium_armor_plate", "中型附加装甲", "装甲", 20, "+4护甲", "-", "-", "强化槽"),
    ("large_armor_plate",  "大型附加装甲", "装甲", 30, "+6护甲", "-", "-", "强化槽"),
]

for row in combat_data:
    ws2.append(["piranport:" + row[0]] + list(row[1:]))

style_rows(ws2)
set_col_widths(ws2, [32, 18, 12, 8, 20, 20, 8, 20])
freeze(ws2)

# ══════════════════════════════════════════════════════════════════════════
# Sheet 3: 航空系统
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("航空系统")
header(ws3, ["注册名", "中文名", "类型", "面板伤害", "航速", "弹药容量", "燃料容量", "负重", "弹药类型", "攻击方式"])

aircraft_data = [
    ("fighter_squadron",       "战斗机编队",     "FIGHTER",       18, 1.8, 64, 1200, 12, "子弹",     "悬停11格，每5tick射1发，panelDmg/8"),
    ("dive_bomber_squadron",   "俯冲轰炸机编队", "DIVE_BOMBER",   24, 1.4, 8,  1200, 16, "航空炸弹", "爬升+18格→俯冲接触<4格投弹，×1.5伤害+50%着火"),
    ("torpedo_bomber_squadron","鱼雷轰炸机编队", "TORPEDO_BOMBER",30, 1.2, 6,  1200, 20, "航空鱼雷", "低飞，距目标20-30格发射鱼雷"),
    ("level_bomber_squadron",  "水平轰炸机编队", "LEVEL_BOMBER",  36, 1.0, 8,  1200, 24, "航空炸弹", "爬升+32格→正上方<3格投弹，×1.5伤害，8发"),
]

for row in aircraft_data:
    ws3.append(["piranport:" + row[0]] + list(row[1:]))

style_rows(ws3)
set_col_widths(ws3, [32, 18, 16, 10, 8, 10, 10, 8, 12, 38])
freeze(ws3)

# ══════════════════════════════════════════════════════════════════════════
# Sheet 4: 食物数值
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("食物数值")
header(ws4, ["注册名", "中文名", "饱食度", "饱和度", "容器", "份数", "Buff效果", "Buff时长(s)"])

food_data = [
    ("toast_bread",         "吐司面包",   15, 18.8, "盘",   3, "—",                              "—"),
    ("naval_baked_beans",   "海军烘豆子",  4,  5.0, "盘",   2, "—",                              "—"),
    ("latiao",              "辣条",        2,  2.5, "盘",   2, "迅捷III",                         90),
    ("mapo_tofu",           "麻婆豆腐",    4,  5.0, "盘",   3, "力量II + 抗火I",                 180),
    ("naval_curry",         "海军咖喱",    5,  6.3, "盘",   3, "夜视I",                           240),
    ("fried_fish_and_chips","炸鱼薯条",    5,  6.3, "盘",   2, "跳跃提升II",                      180),
    ("scone",               "司康饼",      3,  3.8, "盘",   4, "—",                              "—"),
    ("salted_egg_tofu",     "咸蛋拌豆腐",  3,  3.8, "盘",   1, "—",                              "—"),
    ("surstromming",        "鲱鱼罐头",    4,  5.0, "盘",   2, "凋零II(2s)+反胃IV(14s)+力量II",   240),
    ("american_burger",     "美式汉堡",    8, 10.0, "盘",   2, "急迫II",                          180),
    ("hotdog",              "热狗",        4,  5.0, "无",   1, "急迫I",                           180),
    ("pasta",               "意面",        4,  5.0, "盘",   2, "—",                              "—"),
    ("cooked_rice",         "米饭",        5,  6.3, "盘",   2, "—",                              "—"),
    ("beet_blossom",        "甜豆花",      3,  3.8, "碗",   1, "力量I + 水下呼吸I",               180),
    ("miso_soup",           "味噌汤",      6,  7.5, "碗",   2, "—",                              "—"),
]

for row in food_data:
    ws4.append(["piranport:" + row[0]] + list(row[1:]))

style_rows(ws4)
set_col_widths(ws4, [28, 16, 8, 8, 6, 6, 32, 14])
freeze(ws4)

# ══════════════════════════════════════════════════════════════════════════
# Sheet 5: 合成表
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("合成表")
header(ws5, ["产物注册名", "产物中文名", "合成类型", "产量", "材料"])

recipes = [
    # 基础材料
    ("aluminum_ingot",        "铝锭",          "熔炉/高炉",   1,  "矾土矿"),
    ("aluminum_block",        "铝块",          "工作台(有序)", 1,  "铝锭×9（3×3）"),
    ("aluminum_ingot",        "铝锭",          "工作台(无序)", 9,  "铝块×1"),
    ("salt_block",            "盐块",          "工作台(有序)", 1,  "盐×9（3×3）"),
    ("salt",                  "盐",            "工作台(无序)", 9,  "盐块×1"),
    # 核心
    ("small_ship_core",       "小型舰装核心",  "工作台(有序)", 1,  "铝锭×2 + 铁锭×2（2×2交替）"),
    ("medium_ship_core",      "中型舰装核心",  "工作台(有序)", 1,  "铝锭×3 + 铁锭×3（2×3交替）"),
    ("large_ship_core",       "大型舰装核心",  "工作台(有序)", 1,  "铝锭×4 + 铁锭×4 + 铝锭×1（3×3交替）"),
    # 火炮
    ("small_gun",             "小型火炮",      "工作台(有序)", 1,  "铁锭(上)+铝锭(中)+红石(下)"),
    ("medium_gun",            "中型火炮",      "工作台(有序)", 1,  "铁锭×2+铝锭×2+红石(2×3)"),
    ("large_gun",             "大型火炮",      "工作台(有序)", 1,  "铁锭×3+铝锭×3+红石(3×3)"),
    # 炮弹
    ("small_he_shell",        "小口径HE炮弹",  "工作台(有序)", 4,  "铁锭(上)+火药(中)+红石(下)"),
    ("medium_he_shell",       "中口径HE炮弹",  "工作台(有序)", 2,  "铁锭+火药+红石（十字形）"),
    ("large_he_shell",        "大口径HE炮弹",  "工作台(有序)", 1,  "铁锭×6+火药×1+红石×2（3×3）"),
    ("small_ap_shell",        "小口径AP炮弹",  "工作台(有序)", 4,  "铁锭(上)+红石(中)+火药(下)"),
    ("medium_ap_shell",       "中口径AP炮弹",  "工作台(有序)", 2,  "铁锭+红石+火药（十字形）"),
    ("large_ap_shell",        "大口径AP炮弹",  "工作台(有序)", 1,  "铁锭×6+红石×2+火药×1（3×3）"),
    # 鱼雷
    ("torpedo_533mm",         "533mm标准鱼雷", "工作台(有序)", 4,  "铁锭×4+铝锭×2+红石×2+火药×1（3×3）"),
    ("torpedo_610mm",         "610mm氧气鱼雷", "工作台(有序)", 4,  "铁锭×5+铝锭×2+红石×2+火药×1（3×3）"),
    ("twin_torpedo_launcher", "双联装鱼雷发射器","工作台(有序)",1, "铁锭×7+铝锭×1+红石×1（3×3）"),
    ("triple_torpedo_launcher","三联装鱼雷发射器","工作台(有序)",1,"铁锭×7+铝锭×1+红石×1（3×3）"),
    ("quad_torpedo_launcher", "四联装鱼雷发射器","工作台(有序)",1, "铁锭×6+铝锭×2+红石×1（3×3）"),
    # 航弹
    ("aerial_bomb_small",     "小型航弹(旧)",  "工作台(有序)", 4,  "铁锭×4+火药×1（十字形）"),
    ("aerial_bomb_medium",    "中型航弹(旧)",  "工作台(有序)", 2,  "铁锭×8+火药×1（3×3）"),
    ("aerial_torpedo",        "航空鱼雷",      "工作台(有序)", 4,  "铝锭×8+533mm鱼雷×1（3×3）"),
    ("aviation_fuel",         "航空燃料",      "工作台(无序)", 4,  "煤炭×1+桶×1+食用油×1"),
    # 飞机
    ("fighter_squadron",      "战斗机编队",    "工作台(有序)", 1,  "铝锭×4+铁锭×1（十字形）"),
    ("dive_bomber_squadron",  "俯冲轰炸机编队","工作台(有序)", 1,  "铝锭×4+铁锭×2（不规则）"),
    ("torpedo_bomber_squadron","鱼雷轰炸机编队","工作台(有序)",1,  "铝锭×4+铁锭×4+铝锭×1（3×3交替）"),
    ("level_bomber_squadron", "水平轰炸机编队","工作台(有序)", 1,  "铝锭×4+铁锭×4（3×3斜）"),
    # 教程书
    ("guidebook",             "航行手册",      "工作台(有序)", 1,  "铝锭(上)+书(下)"),
    # 石磨配方
    ("flour",                 "面粉",          "石磨",         1,  "小麦×1"),
    ("rice_flour",            "米粉",          "石磨",         1,  "米×1"),
    ("chili_powder",          "辣椒粉",        "石磨",         1,  "辣椒×1+盐×1"),
    ("pork_paste",            "猪肉糜",        "石磨",         1,  "生猪排×1"),
    # 砧板配方
    ("toast_bread_slices",    "吐司面包片",    "砧板(3刀)",    3,  "面包×1"),
    ("bacon",                 "培根",          "砧板(4刀)",    4,  "熟猪排×1"),
    ("sliced_sausage",        "切片香肠",      "砧板(4刀)",    4,  "香肠×1"),
    # 厨锅配方
    ("toast_bread",           "吐司面包",      "厨锅(200t)",   1,  "面粉×2+黄油×1"),
    ("naval_baked_beans",     "海军烘豆子",    "厨锅(200t)",   1,  "大豆×2+番茄×1"),
    ("latiao",                "辣条",          "厨锅(100t)",   1,  "辣椒粉×1+酱油×1"),
    ("mapo_tofu",             "麻婆豆腐",      "厨锅(300t)",   1,  "豆腐×1+辣椒粉×1+酱油×1"),
    ("naval_curry",           "海军咖喱",      "厨锅(300t)",   1,  "土豆×1+胡萝卜×1+熟牛排×1"),
    ("fried_fish_and_chips",  "炸鱼薯条",      "厨锅(200t)",   1,  "鳕鱼×1+土豆×1+食用油×1"),
    ("scone",                 "司康饼",        "厨锅(200t)",   1,  "面粉×1+奶油×1+糖×1"),
    ("salted_egg_tofu",       "咸蛋拌豆腐",    "厨锅(100t)",   1,  "豆腐×1+盐×1"),
    ("beet_blossom",          "甜豆花",        "厨锅(200t)",   1,  "豆腐×1+糖×1"),
    # 酿造配方
    ("yeast",                 "酵母瓶",        "酿造台",        3,  "水瓶×3 + 面粉/米粉/糖"),
    ("brine",                 "盐水",          "酿造台",        3,  "水瓶×3 + 盐"),
    ("miso",                  "味噌",          "酿造台",        3,  "盐水×3 + 大豆"),
    ("soy_sauce",             "酱油",          "酿造台",        3,  "酵母瓶×3 + 大豆"),
    ("vinegar",               "醋",            "酿造台",        3,  "酵母瓶×3 + 米"),
    ("cooking_wine",          "料酒",          "酿造台",        3,  "酵母瓶×3 + 米"),
    ("beer",                  "啤酒",          "酿造台",        3,  "酵母瓶×3 + 小麦"),
]

for row in recipes:
    ws5.append(["piranport:" + row[0]] + list(row[1:]))

style_rows(ws5)
set_col_widths(ws5, [32, 18, 14, 6, 40])
freeze(ws5)

# ── 保存 ──────────────────────────────────────────────────────────────────
out = "/Users/lianran/apps/皮兰港实验/皮兰港_物品数值表.xlsx"
wb.save(out)
print(f"已生成: {out}")
